"""
Excel 출력 모듈.
Sheet 1: 3단비교표 / Sheet 2: 준수여부 / Sheet 3: 누락된 시행령 / Sheet 4: 누락된 시행규칙
"""
import re
import zipfile
from typing import Optional
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# 헤더 스타일
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 짝수 행 배경색
_EVEN_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_ODD_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 컬럼별 기본 너비 (글자 수 기준)
_COL_WIDTHS = {
    "법률 조문":       30,
    "시행령 조문":     30,
    "시행규칙 조문":   30,
    "해당여부":        12,
    "법률 조문번호":   16,
    "법률 조문제목":   28,
    "위임규칙 법령명": 32,
    "위임규칙 조문번호": 16,
    "위임규칙 조문제목": 28,
    "조문번호": 16,
    "조문제목": 32,
}


def _style_sheet(ws, df: pd.DataFrame, center_data: bool = False) -> None:
    """워크시트에 헤더 스타일, 행 교대 색상, 테두리, 열 너비 적용."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COL_WIDTHS.get(col_name, 22)

        # 헤더 셀
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER

    # 데이터 행
    h_align = "center" if center_data else "left"
    for row_idx in range(2, ws.max_row + 1):
        row_fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        if center_data:
            ws.row_dimensions[row_idx].height = 40
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
            cell.border    = _BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _merge_main_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    6컬럼 → 3컬럼: 조문번호 + '\\n' + 조문제목 합치기.
    제목이 없으면 번호만, 번호도 없으면 빈 문자열.
    """
    def combine(num: str, title: str) -> str:
        if not num:
            return ""
        return f"{num}\n{title}" if title else num

    return pd.DataFrame({
        "법률 조문":     [combine(r["법률 조문번호"],     r["법률 조문제목"])     for _, r in df.iterrows()],
        "시행령 조문":   [combine(r["시행령 조문번호"],   r["시행령 조문제목"])   for _, r in df.iterrows()],
        "시행규칙 조문": [combine(r["시행규칙 조문번호"], r["시행규칙 조문제목"]) for _, r in df.iterrows()],
    })


def _add_haedan_dropdown(ws, nrows: int, col: str = "D") -> None:
    """지정 열(해당여부)에 O/X 드롭다운 유효성 검사 추가."""
    dv = DataValidation(
        type="list",
        formula1='"O,X"',
        allow_blank=True,
        showErrorMessage=True,
        error="O 또는 X만 입력할 수 있습니다.",
        errorTitle="입력 오류",
    )
    dv.sqref = f"{col}2:{col}{nrows + 1}"
    ws.add_data_validation(dv)


def _create_junsu_sheet(
    wb,
    nrows1: int,
    columns,
    nrows2: int = 0,
    nrows3: int = 0,
    law_name: str = "",
) -> None:
    """
    '4. 법규준수평가' 시트 생성.
    1행: 문서 제목 / 2행: 법규명 / 3행: 컬럼 헤더 / 4행~: FILTER 수식 데이터
    1. 전체법령(3단) → 2. 독립 시행령 → 3. 독립 시행규칙 순으로
    해당여부 = "O" 인 행을 VSTACK으로 합쳐 출력.
    Excel 365 / 2021 이상 필요.
    """
    ws = wb.create_sheet("4. 법규준수평가")

    # ── 1행: 문서 제목 ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "2026년 안전보건 법규 검토서"
    t.font      = Font(name="맑은 고딕", size=25)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    # ── 2행: 소제목 (법규명) ─────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"법규명 : {law_name}"
    s.font      = Font(name="맑은 고딕", size=20, bold=True)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 38

    # ── 3행: 컬럼 헤더 (No. + 조문 제거) ────────────────────────────────────
    # No. 열 (A열)
    ws.column_dimensions["A"].width = 8
    no_cell = ws.cell(row=3, column=1)
    no_cell.value     = "No."
    no_cell.fill      = _HEADER_FILL
    no_cell.font      = _HEADER_FONT
    no_cell.alignment = _HEADER_ALIGN
    no_cell.border    = _BORDER

    # 데이터 열 (B~D: 법률, 시행령, 시행규칙 / E~G: 추가 입력 열)
    _col_rename = {"법률 조문": "법률", "시행령 조문": "시행령", "시행규칙 조문": "시행규칙"}
    display_cols = [c for c in columns if c != "해당여부"]
    for col_idx, col_name in enumerate(display_cols, start=2):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 18
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = _col_rename.get(col_name, col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER

    # 추가 입력 열 (E~G)
    _extra_cols = [
        (5, "관련 업무 내용", 50),
        (6, "주관부서",       15),
        (7, "평가 결과",      15),
    ]
    for col_idx, col_label, col_width in _extra_cols:
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_width
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = col_label
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER
    ws.row_dimensions[3].height = 30

    ws.freeze_panes = "A4"

    # ── 페이지 설정: 가로 방향 / 열 맞춤 / 3행 반복 인쇄 ────────────────────
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "3:3"  # 2페이지부터 3행(헤더)만 반복; 1행/2행은 1페이지에만 자연 출력

    # ── 페이지 여백 (단위: 인치 / 1.0cm=0.394in, 1.5cm=0.591in) ────────────
    ws.page_margins.top    = 0.394   # 위  1.0 cm
    ws.page_margins.bottom = 0.394   # 아래 1.0 cm
    ws.page_margins.left   = 0.591   # 왼쪽 1.5 cm
    ws.page_margins.right  = 0.591   # 오른쪽 1.5 cm
    ws.page_margins.header = 0.394   # 머리글 1.0 cm
    ws.page_margins.footer = 0.394   # 바닥글 1.0 cm

    # 수식 행수 사전 계산 (데이터 서식·수식·CF 루프에서 공유)
    nrows_max = nrows1 + nrows2 + nrows3

    # ── 데이터 영역 사전 서식: 가운데 정렬 + 줄바꿈 ───────────────────────────
    # 테두리는 아래 조건부 서식으로 처리 (데이터 있는 행에만 적용)
    # 범위를 nrows_max 로 한정 → 여분 행이 인쇄영역을 차지하지 않음
    _da = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(4, 4 + nrows_max):
        for col_idx in range(1, 8):
            ws.cell(row=row_idx, column=col_idx).alignment = _da

    # ── 수식 조립 ────────────────────────────────────────────────────────────
    # 전략: 스필(spill) 수식 대신 행별 독립 수식 INDEX(VSTACK/FILTER, i, col) 사용
    #   → 각 셀이 독립 수식 → Excel이 wrap_text=True 기준으로 행 높이 자동 조절
    # ※ t="array" 패치 필요: FILTER without t="array" → 스칼라 컨텍스트 평가
    #   → 첫 번째 결과만 반환 → i>1 인덱스가 모두 빈 값
    #   → _patch_filter_formula() 에서 각 셀에 t="array" ref="셀주소" 주입
    #
    # _xlfn. 접두사: FILTER, VSTACK, HSTACK 등 Excel 365 신규 함수 필수 표기
    # 빈 셀 처리: openpyxl이 "" → 빈셀 저장 → FILTER가 0 반환
    #             → IF(range=0,"",range) 로 0→"" 선처리
    # 시트2/3 빈 열: IF(LEN(C열)>=0,"","") 로 n×1 "" 배열 생성 후 HSTACK

    fallback3 = '{"","",""}'   # 1×3 공백행 — VSTACK 열수 통일용

    last1  = nrows1 + 1
    range1 = f"'1. 전체법령(3단)'!A2:C{last1}"
    cond1  = f"'1. 전체법령(3단)'!D2:D{last1}=\"O\""
    part1  = (
        f"_xlfn.FILTER("
        f"IF({range1}=0,\"\",{range1}),"
        f"{cond1},{fallback3})"
    )
    parts = [part1]

    if nrows2 > 0:
        last2 = nrows2 + 1
        c2    = f"'2. 독립 시행령'!C2:C{last2}"
        blank = f"IF(LEN({c2})>=0,\"\",\"\")"
        data2 = f"'2. 독립 시행령'!A2:A{last2}&CHAR(10)&'2. 독립 시행령'!B2:B{last2}"
        parts.append(
            f"IFERROR("
            f"_xlfn.FILTER(_xlfn.HSTACK({blank},{data2},{blank}),{c2}=\"O\"),"
            f"{fallback3})"
        )

    if nrows3 > 0:
        last3 = nrows3 + 1
        c3    = f"'3. 독립 시행규칙'!C2:C{last3}"
        blank = f"IF(LEN({c3})>=0,\"\",\"\")"
        data3 = f"'3. 독립 시행규칙'!A2:A{last3}&CHAR(10)&'3. 독립 시행규칙'!B2:B{last3}"
        parts.append(
            f"IFERROR("
            f"_xlfn.FILTER(_xlfn.HSTACK({blank},{blank},{data3}),{c3}=\"O\"),"
            f"{fallback3})"
        )

    vstack_expr = (
        "_xlfn.VSTACK(" + ",".join(parts) + ")"
        if len(parts) > 1 else parts[0]
    )

    # 행별 독립 수식 삽입 (최대 가능 행수만큼)
    # A열: B열에 값이 있을 때만 번호 표시
    # B~D열: INDEX(vstack_expr, i, 열번호) → 단일 값 반환, 스필 없음
    for i, row_idx in enumerate(range(4, 4 + nrows_max), start=1):
        ws.cell(row=row_idx, column=1).value = f'=IF(OR(B{row_idx}<>"",C{row_idx}<>"",D{row_idx}<>""),ROW()-3,"")'
        ws.cell(row=row_idx, column=2).value = f'=IFERROR(INDEX({vstack_expr},{i},1),"")'
        ws.cell(row=row_idx, column=3).value = f'=IFERROR(INDEX({vstack_expr},{i},2),"")'
        ws.cell(row=row_idx, column=4).value = f'=IFERROR(INDEX({vstack_expr},{i},3),"")'

    # ── 조건부 서식: 데이터 있는 행에만 실선 테두리 + 인쇄영역 설정 ───────────
    if nrows_max > 0:
        cf_last_row = 3 + nrows_max
        # B,C,D 중 하나라도 값이 있는 행의 A~G 전체에 테두리 적용
        ws.conditional_formatting.add(
            f'A4:G{cf_last_row}',
            FormulaRule(formula=['OR($B4<>"",$C4<>"",$D4<>"")'], border=_BORDER),
        )
        # 인쇄영역: 제목+헤더(1~3행) + 수식 행수만큼만 명시 (빈 행 제외)
        ws.print_area = f'$A$1:$G${cf_last_row}'


def _patch_filter_formula(output_path: str) -> None:
    """
    저장된 xlsx ZIP을 열어 4. 법규준수평가 시트의 IFERROR(INDEX(_xlfn...) 수식 셀 각각에
    t="array" ref="셀주소" 속성을 추가.

    이유: t="array" 없이는 Excel이 FILTER/VSTACK을 스칼라 컨텍스트로 평가해
    첫 번째 결과만 반환 → INDEX(result, i>1, j) 가 모두 빈 값이 된다.
    openpyxl이 이 속성을 직접 지원하지 않아 XML을 직접 패치한다.
    """
    with zipfile.ZipFile(output_path, 'r') as zf:
        file_map = {n: zf.read(n) for n in zf.namelist()}

    wb_text  = file_map.get('xl/workbook.xml', b'').decode('utf-8')
    rel_text = file_map.get('xl/_rels/workbook.xml.rels', b'').decode('utf-8')

    # 4. 법규준수평가 시트의 rId 추출
    m = re.search(r'<sheet\b[^>]*\bname="4\. 법규준수평가"[^>]*/>', wb_text)
    if not m:
        return
    rid_m = re.search(r'\br:id="([^"]+)"', m.group(0))
    if not rid_m:
        return
    rid = rid_m.group(1)

    # rId → 실제 XML 파일 경로
    sheet_key = None
    for rel_m in re.finditer(r'<Relationship\b[^>]*/>', rel_text):
        tag = rel_m.group(0)
        if f'Id="{rid}"' not in tag:
            continue
        tgt_m = re.search(r'\bTarget="([^"]+)"', tag)
        if tgt_m:
            sheet_key = tgt_m.group(1).lstrip('/')
        break

    if not sheet_key or sheet_key not in file_map:
        return

    sheet_xml = file_map[sheet_key].decode('utf-8')

    # IFERROR(INDEX(_xlfn. 수식을 포함하는 각 셀에 t="array" ref="셀주소" 주입.
    # 패턴: <c r="B5" ...>...<f>IFERROR(INDEX(_xlfn.
    #   → 셀주소(B5)를 추출해 <f t="array" ref="B5"> 로 교체.
    patched = re.sub(
        r'(<c\b[^>]*\br="([^"]+)"[^>]*>(?:(?!</c>).)*?)<f>(IFERROR\(INDEX\(_xlfn\.)',
        lambda mo: f'{mo.group(1)}<f t="array" ref="{mo.group(2)}">{mo.group(3)}',
        sheet_xml,
        flags=re.DOTALL,
    )

    if patched == sheet_xml:
        return

    file_map[sheet_key] = patched.encode('utf-8')
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in file_map.items():
            zf.writestr(name, data)


def export(
    output_path: str,
    main_df: pd.DataFrame,
    missing_enforcement_df: Optional[pd.DataFrame],
    missing_rules_df: Optional[pd.DataFrame],
    law_name: str = "",
) -> None:
    """
    시트 구성 (순서):
      1. 전체법령(3단)   — 법률·시행령·시행규칙 매핑 + 해당여부 드롭다운
      2. 독립 시행령     — 3단비교에 없는 시행령 조문
      3. 독립 시행규칙   — 3단비교에 없는 시행규칙 조문
      4. 법규준수평가    — 해당여부 O 행 자동 연동 (FILTER 수식)
    """
    empty_gap = pd.DataFrame(columns=["조문번호", "조문제목"])

    enf_df  = missing_enforcement_df if missing_enforcement_df is not None else empty_gap
    rule_df = missing_rules_df        if missing_rules_df        is not None else empty_gap

    display_main = _merge_main_df(main_df)
    display_main["해당여부"] = ""  # D열: 사용자가 O/X 직접 선택
    nrows = len(display_main)

    # 시트 2/3: 해당여부 C열 추가
    enf_display  = enf_df.copy();  enf_display["해당여부"]  = ""
    rule_display = rule_df.copy(); rule_display["해당여부"] = ""
    nrows2, nrows3 = len(enf_display), len(rule_display)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        display_main.to_excel(writer, sheet_name="1. 전체법령(3단)", index=False)
        ws_main = writer.sheets["1. 전체법령(3단)"]
        _style_sheet(ws_main, display_main, center_data=True)
        _add_haedan_dropdown(ws_main, nrows)

        enf_display.to_excel(writer, sheet_name="2. 독립 시행령", index=False)
        ws_enf = writer.sheets["2. 독립 시행령"]
        _style_sheet(ws_enf, enf_display, center_data=True)
        if nrows2 > 0:
            _add_haedan_dropdown(ws_enf, nrows2, col="C")

        rule_display.to_excel(writer, sheet_name="3. 독립 시행규칙", index=False)
        ws_rule = writer.sheets["3. 독립 시행규칙"]
        _style_sheet(ws_rule, rule_display, center_data=True)
        if nrows3 > 0:
            _add_haedan_dropdown(ws_rule, nrows3, col="C")

        # 4. 법규준수평가 시트 (FILTER/VSTACK 수식 연동) — 마지막에 생성
        _create_junsu_sheet(writer.book, nrows, display_main.columns.tolist(), nrows2, nrows3, law_name)

    # ExcelWriter가 파일을 닫은 뒤 동적 배열 마커 주입
    _patch_filter_formula(output_path)


def export_simple_list(
    output_path: str,
    articles: list[dict],
) -> None:
    """
    단독 부령용 단일 시트 출력.
    시트: '조문 목록' — 조문번호 | 조문제목 | 해당여부 (O/X 드롭다운)
    """
    df = pd.DataFrame({
        "조문번호": [a["번호"] for a in articles],
        "조문제목": [a["제목"] for a in articles],
        "해당여부": [""] * len(articles),
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="조문 목록", index=False)
        ws = writer.sheets["조문 목록"]
        _style_sheet(ws, df, center_data=True)
        _add_haedan_dropdown(ws, len(df), col="C")
