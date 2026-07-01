"""
기존 법규준수평가표 .xlsm 파일 업데이트.

- 각 법령 시트(2번~)를 최신 API 조문으로 교체
- 기존 D열(해당여부) O/X 보존
- 개정 행: 연노랑(FFF2CC) / 실제 개정된 tier(법·시행령·시행규칙) 셀: 진노랑(FFD966)
- 번호·제목이 같아도 조문 본문(내용해시)이 달라지면 '내용변경'으로 감지
- '변경사항' 시트: tier별 변경 이력 누적
"""
import re
from datetime import datetime
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from src.exporter import (
    _style_sheet, _add_haedan_dropdown, _add_title_row, _set_print_layout,
    _HEADER_FILL, _HEADER_FONT, _HEADER_ALIGN, _BORDER,
)


def _header_row(ws) -> int:
    """헤더('No.')가 위치한 행 탐색. 제목 행 유무(구·신 레이아웃) 모두 대응."""
    for r in (1, 2, 3):
        if str(ws.cell(r, 1).value).strip() == "No.":
            return r
    return 1

_ROW_CHANGE_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 개정 행
_TIER_CHANGE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # 개정된 tier 셀
_HISTORY_SHEET = "변경사항"
_HISTORY_HEADERS = ["업데이트 날짜", "법령명", "구분", "조문번호", "조문명", "변경유형"]

_ART_NUM_RE = re.compile(r'^제\d+조(?:의\d+)?')
_COLS       = ["A", "B", "C"]
_COL_NAMES  = ["법률 조문", "시행령 조문", "시행규칙 조문"]

# tier: (시트 열번호, 새 DataFrame 열이름, 구분 라벨)
_TIERS_STD = [(2, "법률 조문", "법"), (3, "시행령 조문", "시행령"), (4, "시행규칙 조문", "시행규칙")]
_TIERS_ADM = [(2, "조문제목", "조문")]

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─── 셀 파싱 헬퍼 ─────────────────────────────────────────────────────────────

def _parse_art_cell(value: str) -> tuple[str, str]:
    """"제118조\n규제의 재검토" → ("제118조", "규제의 재검토")"""
    if not value:
        return "", ""
    parts = str(value).split("\n", 1)
    num   = parts[0].strip()
    title = parts[1].strip() if len(parts) > 1 else ""
    return num, title


def _primary_key_col(a, b, c) -> tuple[str, str, str]:
    """
    A→B→C 순으로 첫 번째 유효 조문 반환.
    반환: (열 레터, 조문번호, 조문명)
    단독 부령처럼 번호가 "1, 2, 3" 형태여도 허용.
    """
    for col, val in zip(_COLS, [a, b, c]):
        num, title = _parse_art_cell(str(val) if val else "")
        if num:
            return col, num, title
    return "", "", ""


# ─── O/X 보존 ────────────────────────────────────────────────────────────────

def _read_ox_map(ws, hdr: int = 1) -> tuple[dict, dict]:
    """
    열+번호 조합 키로 O/X 추출.
    반환: (ox_by_num={(col,num):ox}, ox_by_title={(col,title):ox})
    """
    ox_by_num:   dict[tuple, str] = {}
    ox_by_title: dict[tuple, str] = {}

    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if len(row) < 5:
            continue
        a, b, c, d = row[1], row[2], row[3], row[4]
        ox = str(d).strip() if d else ""
        if not ox:
            continue
        col, num, title = _primary_key_col(a, b, c)
        if col and num:
            ox_by_num.setdefault((col, num), ox)
        if col and title:
            ox_by_title.setdefault((col, title), ox)

    return ox_by_num, ox_by_title


def _row_triple(row, is_admrul: bool) -> tuple[str, str, str]:
    """새 DataFrame 행에서 키 산출용 (a, b, c) 추출.
    행정규칙은 조문제목(B)만 키로 사용."""
    if is_admrul:
        return str(row.get("조문제목", "") or ""), "", ""
    return row["법률 조문"], row["시행령 조문"], row["시행규칙 조문"]


def _apply_ox(new_df: pd.DataFrame,
              ox_by_num: dict, ox_by_title: dict,
              is_admrul: bool = False) -> pd.DataFrame:
    """새 DataFrame 각 행에 기존 O/X 매핑."""
    result = new_df.copy()
    for idx, row in result.iterrows():
        col, num, title = _primary_key_col(*_row_triple(row, is_admrul))
        ox = (ox_by_num.get((col, num)) or
              ox_by_title.get((col, title)) or "")
        result.at[idx, "해당여부"] = ox
    return result


# ─── 변경 감지 ────────────────────────────────────────────────────────────────

def _read_old_articles(ws, hdr: int = 1) -> tuple[dict, dict]:
    """기존 시트에서 tier별 제목·내용해시 반환.
    반환: (old_by_tier={시트열: {번호: {제목들}}}, old_hash={시트열: {번호: 내용해시}})
    법(B)·시행령(C)·시행규칙(D)을 각각 독립 수집. 내용해시는 숨김 G열(7번째)의
    '법|시행령|시행규칙' 해시에서 tier별로 파싱. (행정규칙은 조문제목(B) 1개 tier.)"""
    is_adm = str(ws.cell(hdr, 2).value).strip() == "조문제목"
    tiers  = _TIERS_ADM if is_adm else _TIERS_STD
    old:  dict[int, dict[str, set]] = {sc: {} for sc, _, _ in tiers}
    oldh: dict[int, dict[str, str]] = {sc: {} for sc, _, _ in tiers}
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        gval = str(row[6]) if len(row) >= 7 and row[6] else ""
        segs = gval.split("|")
        for i, (sc, _, _) in enumerate(tiers):
            val = row[sc - 1] if len(row) >= sc else None
            num, title = _parse_art_cell(str(val) if val else "")
            if num:
                old[sc].setdefault(num, set()).add(title)
                if i < len(segs) and segs[i]:
                    oldh[sc].setdefault(num, segs[i])
    return old, oldh


def _detect_changes(old_by_tier: dict, old_hash: dict,
                    new_df: pd.DataFrame,
                    is_admrul: bool = False) -> tuple[list, list]:
    """
    tier별 조문 비교로 개정 항목을 감지한다.
    번호·제목이 같아도 내용해시가 달라지면 '내용변경'으로 감지한다.
    반환: (row_orange, changes)
      row_orange : 각 새 행마다 개정된 tier 시트열번호 집합
      changes    : [{"구분","조문번호","조문명","변경유형"}]
                   (신설/변경/내용변경/삭제, tier별·중복제거)
    """
    tiers = _TIERS_ADM if is_admrul else _TIERS_STD

    # tier별 (제목 → 번호) 역방향 맵
    old_by_title = {sc: {} for sc in old_by_tier}
    for sc, m in old_by_tier.items():
        for num, titles in m.items():
            for t in titles:
                old_by_title[sc].setdefault(t, num)

    matched = {sc: set() for sc in old_by_tier}
    row_orange: list = []
    changes:    list = []
    seen:       set  = set()

    def _add(label, num, title, ctype):
        k = (label, num, title, ctype)
        if k not in seen:
            seen.add(k)
            changes.append({"구분": label, "조문번호": num,
                            "조문명": title, "변경유형": ctype})

    for _, row in new_df.iterrows():
        new_segs = str(row.get("내용해시", "") or "").split("|")
        oranges: set = set()
        for i, (sc, dfcol, label) in enumerate(tiers):
            num, title = _parse_art_cell(str(row.get(dfcol, "") or ""))
            if not num:
                continue
            oldm = old_by_tier.get(sc, {})
            newh = new_segs[i] if i < len(new_segs) else ""
            if num in oldm:
                matched[sc].add(num)
                if title in oldm[num]:
                    oh = old_hash.get(sc, {}).get(num)
                    # 번호·제목 동일 → 내용해시로 본문 변경 확인 (옛 해시 있을 때만)
                    status = "내용변경" if (oh and newh and newh != oh) else "동일"
                else:
                    status = "변경"                        # 제목 변경
            elif title and title in old_by_title.get(sc, {}):
                matched[sc].add(old_by_title[sc][title])    # 번호 이동
                status = "변경"
            else:
                status = "신설"
            if status != "동일":
                oranges.add(sc)
                _add(label, num, title, status)
        row_orange.append(oranges)

    # 삭제: 옛 조문 중 새 데이터에서 매칭되지 않은 것
    for sc, dfcol, label in tiers:
        for num, titles in old_by_tier.get(sc, {}).items():
            if num not in matched[sc]:
                for t in titles:
                    _add(label, num, t, "삭제")

    return row_orange, changes


# ─── 시트 재작성 ─────────────────────────────────────────────────────────────

def _rewrite_sheet(ws, new_df: pd.DataFrame, row_orange: list,
                   law_name: str, is_admrul: bool = False) -> None:
    """기존 내용 전체 삭제 후 제목행+헤더+데이터 재작성.
    개정 행은 A~D 연노랑, 실제 개정된 tier 셀은 진노랑.
    레이아웃: 1행=제목 / 2행=헤더 / 3행~=데이터 (구 레이아웃 파일도 신 레이아웃으로 정규화)."""
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    ncols = len(new_df.columns)
    _add_title_row(ws, law_name, ncols)                    # 1행: 제목
    for col_idx, col_name in enumerate(new_df.columns, start=1):
        ws.cell(row=2, column=col_idx, value=col_name)     # 2행: 헤더

    for _, row in new_df.iterrows():                        # 3행~: 데이터
        if is_admrul:
            ws.append([
                row["No."], row.get("조문제목", ""), row.get("조문내용", ""),
                "", row["해당여부"], row.get("하위조문내용", ""), row.get("내용해시", ""),
            ])
        else:
            ws.append([
                row["No."], row["법률 조문"], row["시행령 조문"],
                row["시행규칙 조문"], row["해당여부"], row.get("하위조문내용", ""), row.get("내용해시", ""),
            ])

    _style_sheet(ws, new_df, header_row=2)

    for row_idx, oranges in enumerate(row_orange, start=3):
        if not oranges:
            continue
        for col_idx in range(1, 5):                     # 개정 행: A~D 연노랑
            ws.cell(row=row_idx, column=col_idx).fill = _ROW_CHANGE_FILL
        for sc in oranges:                              # 실제 개정된 tier 셀: 진노랑
            ws.cell(row=row_idx, column=sc).fill = _TIER_CHANGE_FILL

    _add_haedan_dropdown(ws, len(new_df), start_row=3)
    _set_print_layout(ws)
    ws.column_dimensions["F"].hidden = True  # 하위조문내용(VBA 전용) 숨김
    ws.column_dimensions["G"].hidden = True  # 내용해시(내용변경 감지용) 숨김
    if is_admrul:
        ws.column_dimensions["D"].hidden = True  # 미사용 placeholder 열
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 65  # F열 장문에 따른 행높이 자동팽창 방지(고정 65)


# ─── 변경사항 시트 ────────────────────────────────────────────────────────────

def _ensure_history_sheet(wb):
    """변경사항 시트 없으면 생성, 있으면 반환.
    구(舊) 스키마(구분 열 없음)면 C열에 '구분'을 삽입해 마이그레이션한다."""
    from openpyxl.utils import get_column_letter
    if _HISTORY_SHEET in wb.sheetnames:
        ws = wb[_HISTORY_SHEET]
        if str(ws.cell(1, 3).value).strip() != "구분":
            ws.insert_cols(3)
            cell = ws.cell(row=1, column=3, value="구분")
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border    = _BORDER
            ws.column_dimensions["C"].width = 10
        return ws

    ws = wb.create_sheet(_HISTORY_SHEET)
    ws.append(_HISTORY_HEADERS)
    widths = [14, 36, 10, 14, 32, 10]
    for col_idx, (h, w) in enumerate(zip(_HISTORY_HEADERS, widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    return ws


def _append_history(ws_hist, today: str,
                    law_name: str, changes: list) -> None:
    """변경사항 시트에 이력 행 추가."""
    for change in changes:
        ws_hist.append([
            today,
            law_name,
            change.get("구분", ""),
            change["조문번호"],
            change["조문명"],
            change["변경유형"],
        ])
        row_idx = ws_hist.max_row
        for col_idx in range(1, 7):
            cell = ws_hist.cell(row=row_idx, column=col_idx)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = _BORDER


# ─── 진입점 ───────────────────────────────────────────────────────────────────

def update_file(
    xlsx_path: str,
    process_fn: Callable[[str], tuple[str, pd.DataFrame] | None],
) -> None:
    """
    xlsx_path:  업데이트할 .xlsm / .xlsx 경로
    process_fn: law_name → (actual_name, combined_df) | None
    """
    print(f"\n파일 로드 중: {xlsx_path}")
    wb   = load_workbook(xlsx_path, keep_vba=True)
    today = datetime.now().strftime("%Y/%m/%d")

    updated, skipped = 0, 0
    all_changes: list[tuple[str, dict]] = []   # (law_name, change_dict)

    for sheet in wb.worksheets:
        title = sheet.title
        if ". " not in title:
            continue
        law_name = title.split(". ", 1)[1]
        if law_name in ("법규준수평가", _HISTORY_SHEET):
            continue

        print(f"\n  [{law_name}] 처리 중...")

        hdr = _header_row(sheet)
        ox_by_num, ox_by_title = _read_ox_map(sheet, hdr)
        old_by_tier, old_hash  = _read_old_articles(sheet, hdr)
        preserved = sum(1 for v in ox_by_num.values() if v == "O")
        print(f"    기존 O 표시 조문: {preserved}개 보존 예정")

        result = process_fn(law_name)
        if result is None:
            print(f"    건너뜀 (API 조회 실패)")
            skipped += 1
            continue

        _, new_df = result
        is_admrul = "조문제목" in new_df.columns
        new_df = _apply_ox(new_df, ox_by_num, ox_by_title, is_admrul)

        row_orange, changes = _detect_changes(old_by_tier, old_hash, new_df, is_admrul)

        cnt = {t: sum(1 for c in changes if c["변경유형"] == t)
               for t in ("신설", "변경", "내용변경", "삭제")}
        print(f"    완료 ({len(new_df)}행) — 신설 {cnt['신설']}, 변경 {cnt['변경']}, "
              f"내용변경 {cnt['내용변경']}, 삭제 {cnt['삭제']}")

        _rewrite_sheet(sheet, new_df, row_orange, law_name, is_admrul)
        updated += 1

        # 이력 수집 (tier별)
        for ch in changes:
            all_changes.append((law_name, ch))

    # 변경사항 시트 갱신
    if all_changes:
        ws_hist = _ensure_history_sheet(wb)
        # 법령명별로 묶어서 기록
        current_law, batch = "", []
        for law_name, change in all_changes + [("__END__", None)]:
            if law_name != current_law:
                if batch:
                    _append_history(ws_hist, today, current_law, batch)
                current_law, batch = law_name, []
            if change:
                batch.append(change)
        print(f"\n변경사항 시트: {len(all_changes)}건 기록")
    else:
        print("\n변경사항 없음 — 이력 미기록")

    wb.save(xlsx_path)
    print(f"\n{'─'*50}")
    print(f"업데이트 완료: {updated}개 시트 갱신, {skipped}개 건너뜀")
    print(f"저장: {xlsx_path}")
    print()
    print("※ '1. 법규준수평가' 시트를 갱신하려면 파일을 열고")
    print("  임의 조문의 해당여부(O/X)를 변경하면 자동 반영됩니다.")
