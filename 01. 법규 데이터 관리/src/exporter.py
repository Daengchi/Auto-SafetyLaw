"""
Excel 출력 모듈.
Sheet 1: 3단비교표 / Sheet 2: 준수여부 / Sheet 3: 누락된 시행령 / Sheet 4: 누락된 시행규칙
"""
import os
import zipfile
from typing import Optional
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# 헤더 스타일
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 짝수 행 배경색
_EVEN_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_ODD_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 컬럼별 기본 너비 (글자 수 기준)
_COL_WIDTHS = {
    "No.":             6,
    "법률 조문":       35,
    "시행령 조문":     35,
    "시행규칙 조문":   35,
    "해당여부":        12,
    "법률 조문번호":   16,
    "법률 조문제목":   28,
    "위임규칙 법령명": 32,
    "위임규칙 조문번호": 16,
    "위임규칙 조문제목": 28,
    "조문번호": 16,
    "조문제목": 32,
    "조문내용": 70,
}


def _style_sheet(ws, df: pd.DataFrame, center_data: bool = False) -> None:
    """워크시트에 헤더 스타일, 행 교대 색상, 테두리, 열 너비 적용."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COL_WIDTHS.get(col_name, 22)

        # 헤더 셀
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER

    # 데이터 행
    h_align = "center" if center_data else "left"
    for row_idx in range(2, ws.max_row + 1):
        row_fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        if center_data:
            max_lines = 1
            for ci in range(1, len(df.columns) + 1):
                val = ws.cell(row=row_idx, column=ci).value
                if val:
                    max_lines = max(max_lines, str(val).count('\n') + 1)
            ws.row_dimensions[row_idx].height = max(20, max_lines * 16)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = row_fill
            cell.font      = Font(size=10)
            cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
            cell.border    = _BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _merge_main_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    6컬럼 → 3컬럼: 조문번호 + '\\n' + 조문제목 합치기.
    제목이 없으면 번호만, 번호도 없으면 빈 문자열.
    """
    def combine(num: str, title: str) -> str:
        if not num:
            return ""
        return f"{num}\n{title}" if title else num

    return pd.DataFrame({
        "법률 조문":     [combine(r["법률 조문번호"],     r["법률 조문제목"])     for _, r in df.iterrows()],
        "시행령 조문":   [combine(r["시행령 조문번호"],   r["시행령 조문제목"])   for _, r in df.iterrows()],
        "시행규칙 조문": [combine(r["시행규칙 조문번호"], r["시행규칙 조문제목"]) for _, r in df.iterrows()],
    })


def _build_combined_df(
    main_df: pd.DataFrame,
    enf_df: pd.DataFrame,
    rule_df: pd.DataFrame,
) -> pd.DataFrame:
    """3단법령 + 단독 시행령(B열) + 단독 시행규칙(C열) → 단일 DataFrame."""
    def _art_cell(row) -> str:
        num   = str(row.get("조문번호", "") or "")
        title = str(row.get("조문제목", "") or "")
        return f"{num}\n{title}" if title else num

    merged = _merge_main_df(main_df)
    main_rows = [
        {"법률 조문": r["법률 조문"], "시행령 조문": r["시행령 조문"],
         "시행규칙 조문": r["시행규칙 조문"], "해당여부": ""}
        for _, r in merged.iterrows()
    ]
    enf_rows = [
        {"법률 조문": "", "시행령 조문": _art_cell(r), "시행규칙 조문": "", "해당여부": ""}
        for _, r in enf_df.iterrows()
    ]
    rul_rows = [
        {"법률 조문": "", "시행령 조문": "", "시행규칙 조문": _art_cell(r), "해당여부": ""}
        for _, r in rule_df.iterrows()
    ]
    all_rows = main_rows + enf_rows + rul_rows
    for i, row in enumerate(all_rows, start=1):
        row["No."] = i
    return pd.DataFrame(
        all_rows,
        columns=["No.", "법률 조문", "시행령 조문", "시행규칙 조문", "해당여부"],
    )


def _build_admrul_df(articles: list[dict]) -> pd.DataFrame:
    """
    행정규칙 조문 → 단일 DataFrame.
    3단 위임구조가 없으므로 B=조문제목, C=조문내용 으로 구성.
    해당여부는 E열(5번째)에 유지해야 VBA가 동작하므로 D열(시행규칙 조문)은
    빈 placeholder 로 두고 export 시 숨긴다.
    """
    rows = [
        {"조문제목": f"{a['번호']}\n{a['제목']}" if a.get("제목") else a["번호"],
         "조문내용": a.get("내용", ""),
         "시행규칙 조문": "", "해당여부": ""}
        for a in articles
    ]
    for i, row in enumerate(rows, start=1):
        row["No."] = i
    return pd.DataFrame(
        rows,
        columns=["No.", "조문제목", "조문내용", "시행규칙 조문", "해당여부"],
    )


def _is_admrul_df(df: pd.DataFrame) -> bool:
    """행정규칙 시트 여부 판별 (조문제목 컬럼 보유)."""
    return "조문제목" in df.columns


def _add_haedan_dropdown(ws, nrows: int, col: str = "E") -> None:
    """지정 열(해당여부)에 O/X 드롭다운 유효성 검사 추가."""
    dv = DataValidation(
        type="list",
        formula1='"O,X"',
        allow_blank=True,
        showErrorMessage=True,
        error="O 또는 X만 입력할 수 있습니다.",
        errorTitle="입력 오류",
    )
    dv.sqref = f"{col}2:{col}{nrows + 1}"
    ws.add_data_validation(dv)


def _create_junsu_sheet(wb) -> None:
    """
    '1. 법규준수평가' 시트 생성.
    1행: 문서 제목 / 2행: 빈 행 / 3행: 컬럼 헤더 / 4행~: VBA가 O 행 삽입
    모든 법령 시트(Sheet 2+)의 해당여부 D열 O 입력 시 VBA가 이 시트에 행 추가.
    """
    ws = wb.create_sheet("1. 법규준수평가")

    # ── 1행: 문서 제목 ──────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = "2026년 안전보건 법규 검토서"
    t.font      = Font(name="맑은 고딕", size=25)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    # ── 2행: 빈 행 (행 구조 유지용 — VBA HDR_ROW=3, DAT_ROW=4 고정) ─────────
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 10

    # ── 3행: 컬럼 헤더 ───────────────────────────────────────────────────────
    # No. 열 (A열)
    ws.column_dimensions["A"].width = 8
    no_cell = ws.cell(row=3, column=1)
    no_cell.value     = "No."
    no_cell.fill      = _HEADER_FILL
    no_cell.font      = _HEADER_FONT
    no_cell.alignment = _HEADER_ALIGN
    no_cell.border    = _BORDER

    # 데이터 열 (B~D: 법률, 시행령, 시행규칙 / E~G: 추가 입력 열)
    _col_rename = {"법률 조문": "법률", "시행령 조문": "시행령", "시행규칙 조문": "시행규칙"}
    display_cols = ["법률 조문", "시행령 조문", "시행규칙 조문"]
    for col_idx, col_name in enumerate(display_cols, start=2):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 18
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = _col_rename.get(col_name, col_name)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER

    # 추가 입력 열 (E~I)
    _extra_cols = [
        (5, "주요 내용",    50),
        (6, "주관부서",     15),
        (7, "필요업무내용", 50),
        (8, "준수평가",     15),
        (9, "비고",        15),
    ]
    for col_idx, col_label, col_width in _extra_cols:
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_width
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = col_label
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER
    ws.row_dimensions[3].height = 30

    ws.freeze_panes = "A4"

    # ── 페이지 설정: 가로 방향 / 열 맞춤 / 3행 반복 인쇄 ────────────────────
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "3:3"  # 2페이지부터 3행(헤더)만 반복; 1행/2행은 1페이지에만 자연 출력

    # ── 페이지 여백 (단위: 인치 / 1.0cm=0.394in, 1.5cm=0.591in) ────────────
    ws.page_margins.top    = 0.394   # 위  1.0 cm
    ws.page_margins.bottom = 0.394   # 아래 1.0 cm
    ws.page_margins.left   = 0.591   # 왼쪽 1.5 cm
    ws.page_margins.right  = 0.591   # 오른쪽 1.5 cm
    ws.page_margins.header = 0.394   # 머리글 1.0 cm
    ws.page_margins.footer = 0.394   # 바닥글 1.0 cm

    # ── J·K열: VBA 전용 숨김 열 ─────────────────────────────────────────────
    ws.column_dimensions["J"].hidden = True  # 조문키
    ws.column_dimensions["K"].hidden = True  # 정렬키

    # ── 4행: VBA 서식 참조용 템플릿 행 (값 없음, 서식만) ──────────────────────
    # VBA가 ws4.Rows(4).Copy → PasteSpecial xlPasteFormats 로 신규 행 서식 복사
    _da = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, 10):
        cell = ws.cell(row=4, column=col_idx)
        cell.alignment = _da
        cell.fill = _ODD_FILL


def _inject_vba(xlsx_path: str, xlsm_path: str) -> None:
    """
    xlsx → xlsm 변환 (ZIP 후처리).
    src/template.xlsm 에서 vbaProject.bin 을 추출해 주입하고,
    [Content_Types].xml 및 workbook.xml.rels 를 xlsm 형식으로 수정한다.
    """
    template_path = os.path.join(os.path.dirname(__file__), "template.xlsm")
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"template.xlsm 을 찾을 수 없습니다: {template_path}\n"
            "tools/make_template.py 를 실행하거나 수동으로 생성하세요."
        )

    with zipfile.ZipFile(template_path, 'r') as tz:
        vba_data = tz.read("xl/vbaProject.bin")
        template_wb_xml = tz.read("xl/workbook.xml").decode("utf-8")

    # template workbook.xml 에서 codeName 추출 (ThisWorkbook 이벤트 바인딩용)
    _codename = None
    _pr_idx = template_wb_xml.find('<workbookPr')
    if _pr_idx != -1:
        _pr_end = template_wb_xml.find('>', _pr_idx)
        _pr_tag = template_wb_xml[_pr_idx:_pr_end]
        _cn_idx = _pr_tag.find('codeName="')
        if _cn_idx != -1:
            _cn_start = _cn_idx + len('codeName="')
            _cn_end = _pr_tag.index('"', _cn_start)
            _codename = _pr_tag[_cn_start:_cn_end]

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        file_map = {n: zf.read(n) for n in zf.namelist()}

    # [Content_Types].xml: 워크북 타입을 xlsm 으로 변경 + vbaProject 항목 추가
    ct = file_map["[Content_Types].xml"].decode("utf-8")
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    ct = ct.replace(
        "</Types>",
        '<Override PartName="/xl/vbaProject.bin"'
        ' ContentType="application/vnd.ms-office.vbaProject"/></Types>',
    )
    file_map["[Content_Types].xml"] = ct.encode("utf-8")

    # xl/_rels/workbook.xml.rels: vbaProject 관계 추가
    rels = file_map["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rels = rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdVBA"'
        ' Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject"'
        ' Target="vbaProject.bin"/></Relationships>',
    )
    file_map["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    # xl/workbook.xml: codeName 주입 (ThisWorkbook 이벤트 핸들러 바인딩)
    if _codename:
        wb_xml = file_map["xl/workbook.xml"].decode("utf-8")
        if "<workbookPr/>" in wb_xml:
            wb_xml = wb_xml.replace("<workbookPr/>",
                f'<workbookPr codeName="{_codename}"/>', 1)
        elif "<workbookPr " in wb_xml:
            idx = wb_xml.index("<workbookPr ")
            end = wb_xml.index(">", idx)
            if 'codeName=' not in wb_xml[idx:end]:
                wb_xml = wb_xml.replace("<workbookPr ",
                    f'<workbookPr codeName="{_codename}" ', 1)
        file_map["xl/workbook.xml"] = wb_xml.encode("utf-8")

    file_map["xl/vbaProject.bin"] = vba_data

    with zipfile.ZipFile(xlsm_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in file_map.items():
            zf.writestr(name, data)


def export(
    output_path: str,
    main_df: pd.DataFrame,
    missing_enforcement_df: Optional[pd.DataFrame],
    missing_rules_df: Optional[pd.DataFrame],
    law_name: str = "",
) -> None:
    """
    시트 구성 (순서):
      1. 전체법령(3단)   — 법률·시행령·시행규칙 매핑 + 해당여부 드롭다운
      2. 독립 시행령     — 3단비교에 없는 시행령 조문
      3. 독립 시행규칙   — 3단비교에 없는 시행규칙 조문
      4. 법규준수평가    — VBA Workbook_SheetChange 이벤트로 행 삽입/삭제
    출력 형식: .xlsm (매크로 사용 통합 문서)
    """
    empty_gap = pd.DataFrame(columns=["조문번호", "조문제목"])

    enf_df  = missing_enforcement_df if missing_enforcement_df is not None else empty_gap
    rule_df = missing_rules_df        if missing_rules_df        is not None else empty_gap

    combined_df = _build_combined_df(main_df, enf_df, rule_df)
    nrows = len(combined_df)

    # 데이터를 임시 xlsx 로 먼저 저장, 이후 VBA 주입해 xlsm 으로 변환
    xlsx_tmp = output_path[:-5] + ".xlsx"  # .xlsm → .xlsx

    with pd.ExcelWriter(xlsx_tmp, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="1. 전체법령", index=False)
        ws_main = writer.sheets["1. 전체법령"]
        _style_sheet(ws_main, combined_df, center_data=True)
        _add_haedan_dropdown(ws_main, nrows)

        # 2. 법규준수평가 시트 (VBA 대기용 헤더+템플릿행만 생성)
        _create_junsu_sheet(writer.book)

    # xlsx → xlsm: src/template.xlsm 의 vbaProject.bin 주입
    _inject_vba(xlsx_tmp, output_path)
    os.remove(xlsx_tmp)


def export_multi(
    output_path: str,
    laws: list[tuple[str, pd.DataFrame]],
) -> None:
    """
    Sheet 1: '1. 법규준수평가' (VBA 대기용)
    Sheet 2+: 법령별 데이터 시트 ('2. 법령명', '3. 법령명', ...)
    출력 형식: .xlsm
    """
    xlsx_tmp = output_path[:-5] + ".xlsx"

    with pd.ExcelWriter(xlsx_tmp, engine="openpyxl") as writer:
        # Sheet 2+: 법령별 데이터 (먼저 생성)
        for idx, (law_name, combined_df) in enumerate(laws, start=2):
            sheet_name = f"{idx}. {law_name}"[:31]  # Excel 시트명 31자 제한
            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_sheet(ws, combined_df, center_data=True)
            _add_haedan_dropdown(ws, len(combined_df))
            if _is_admrul_df(combined_df):
                ws.column_dimensions["D"].hidden = True  # 미사용 placeholder 열

        # Sheet 1: 법규준수평가 — 생성 후 맨 앞으로 이동
        _create_junsu_sheet(writer.book)
        wb = writer.book
        wb.move_sheet(wb["1. 법규준수평가"], offset=-(len(wb.sheetnames) - 1))

    _inject_vba(xlsx_tmp, output_path)
    os.remove(xlsx_tmp)


def export_simple_list(
    output_path: str,
    articles: list[dict],
) -> None:
    """
    단독 부령용 단일 시트 출력.
    시트: '조문 목록' — 조문번호 | 조문제목 | 해당여부 (O/X 드롭다운)
    """
    df = pd.DataFrame({
        "조문번호": [a["번호"] for a in articles],
        "조문제목": [a["제목"] for a in articles],
        "해당여부": [""] * len(articles),
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="조문 목록", index=False)
        ws = writer.sheets["조문 목록"]
        _style_sheet(ws, df, center_data=True)
        _add_haedan_dropdown(ws, len(df), col="C")
